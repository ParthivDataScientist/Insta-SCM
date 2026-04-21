"""
Shipments API endpoints - HTTP layer only.
All business logic lives in app.services.shipment_service.
"""
import io
import logging
from datetime import datetime, timezone
from typing import List, Optional

import pandas as pd
from fastapi import APIRouter, Depends, File, HTTPException, Path, Query, UploadFile
from fastapi.responses import StreamingResponse
from fastapi.concurrency import run_in_threadpool
from pydantic import BaseModel
from sqlmodel import Session, select

from app.api.middleware.dhl_validation import validate_dhl_awb_or_400
from app.core.security import verify_api_key
from app.db.session import get_session
from app.models.dashboard_project import DashboardProject
from app.models.shipment import Shipment
from app.schemas.shipment import MPSDetailResponse, ShipmentResponse
from app.services.carrier_detection import detect_carrier
from app.services.dhl_validation import DHL_AWB_FORMAT_ERROR
from app.services.shipment_service import get_stats, track_and_save, preview_track, refresh_tracked_shipments
from app.services.dhl import DHLService

logger = logging.getLogger(__name__)
router = APIRouter()

MAX_EXCEL_FILE_SIZE = 5 * 1024 * 1024  # 5 MB


class TrackRequest(BaseModel):
    """Request body for tracking a new shipment."""
    recipient: Optional[str] = None
    shipment_name: Optional[str] = None
    destination: Optional[str] = None
    show_date: Optional[str] = None
    exhibition_name: Optional[str] = None
    cs: Optional[str] = None
    no_of_box: Optional[str] = None
    project_id: Optional[int] = None


class BatchRequest(BaseModel):
    """Request body for batch operations."""
    shipment_ids: List[int]
    archive: Optional[bool] = None


class RefreshRequest(BaseModel):
    """Request body for re-syncing saved shipment records."""
    shipment_ids: Optional[List[int]] = None
    include_children: bool = False

class SheetRow(BaseModel):
    row_number: Optional[int] = None
    ship_to_location: Optional[str] = None
    client_name: Optional[str] = None
    booking_date: Optional[str] = None
    show_date: Optional[str] = None
    show_city: Optional[str] = None
    cs_type: Optional[str] = None
    no_of_box: Optional[str] = None
    courier: Optional[str] = None
    master_awb: Optional[str] = None
    child_awb: Optional[str] = None
    remarks: Optional[str] = None

class WebhookPayload(BaseModel):
    rows: List[SheetRow]
    track_live: bool = False


EMPTY_TRACKING_TOKENS = {"", "nan", "none", "null", "-", "n/a", "na"}


def _normalize_tracking_cell(value: Optional[str]) -> str:
    token = str(value or "").strip()
    if token.lower() in EMPTY_TRACKING_TOKENS:
        return ""
    if token.endswith(".0") and token[:-2].isdigit():
        token = token[:-2]
    return token.upper()


def _resolve_tracking_row(
    *,
    master_awb: Optional[str],
    child_awb: Optional[str],
    legacy_tracking: Optional[str] = None,
    last_master_awb: Optional[str] = None,
) -> tuple[Optional[str], Optional[str], bool, Optional[str]]:
    """
    Resolve the row tracking number and MPS relationship.

    Returns:
        (tracking_number, master_tracking_number, is_master, next_last_master_awb)
    """
    master = _normalize_tracking_cell(master_awb)
    child = _normalize_tracking_cell(child_awb)
    legacy = _normalize_tracking_cell(legacy_tracking)

    if master and child:
        if master == child:
            return master, None, True, master
        # Google Sheet flow can provide both columns for child rows:
        # master_awb is the parent, child_awb is the parcel being tracked.
        return child, master, False, master

    if master:
        return master, None, True, master

    if child:
        return child, _normalize_tracking_cell(last_master_awb), False, _normalize_tracking_cell(last_master_awb)

    if legacy:
        return legacy, None, False, _normalize_tracking_cell(last_master_awb)

    return None, None, False, _normalize_tracking_cell(last_master_awb)


def _serialize_shipment(db: Session, shipment: Shipment) -> ShipmentResponse:
    project = db.get(DashboardProject, shipment.project_id) if shipment.project_id else None
    payload = ShipmentResponse.model_validate(shipment).model_dump()
    payload["project_name"] = project.project_name if project else None
    payload["project_client_name"] = (
        project.client_relationship.name
        if project and project.client_relationship
        else None
    )
    return ShipmentResponse(**payload)


def _validate_project_reference(db: Session, project_id: Optional[int]) -> Optional[DashboardProject]:
    if project_id is None:
        return None

    project = db.get(DashboardProject, project_id)
    if not project:
        raise HTTPException(status_code=404, detail="Linked project not found")
    return project


# ---------------------------------------------------------------------------
# ENDPOINTS
# ---------------------------------------------------------------------------




@router.get("/track/{tracking_number}/preview")
def preview_shipment(
    tracking_number: str = Path(
        ...,
        min_length=8,
        max_length=50,
        description="Carrier tracking number to preview (no DB save)",
    ),
    master_tracking_number: Optional[str] = Query(
        default=None,
        description="Optional master tracking number hint for child-piece lookups.",
    ),
    db: Session = Depends(get_session),
    _key: str = Depends(verify_api_key),
):
    """Fetch live tracking data for a tracking number WITHOUT saving to the database."""
    result = preview_track(
        tracking_number.upper(),
        db=db,
        master_tracking_number=(master_tracking_number or "").upper() or None,
    )
    if "error" in result:
        raise HTTPException(status_code=400, detail=result["error"])
    return result


@router.get("/dhl/track/{awb}/preview")
def preview_dhl_shipment(
    awb: str = Path(
        ...,
        min_length=10,
        max_length=20,
        description="DHL Express India AWB (10 digits)",
    ),
    _key: str = Depends(verify_api_key),
):
    """
    DHL-only preview endpoint with strict AWB validation and isolated DHL provider flow.
    """
    normalized_awb = validate_dhl_awb_or_400(awb)
    result = DHLService().track(normalized_awb)
    if "error" in result:
        raise HTTPException(status_code=400, detail=result["error"])
    return {
        "carrier": "DHL",
        "tracking_number": normalized_awb,
        "current_status": result.get("current_status"),
        "estimated_delivery": result.get("estimated_delivery"),
        "last_location": result.get("last_location"),
    }


@router.post("/track/{tracking_number}", status_code=201)
def track_shipment(
    tracking_number: str = Path(
        ...,
        min_length=8,
        max_length=50,
        pattern=r"^[A-Z0-9]+$",
        description="Carrier tracking number (uppercase alphanumeric, 8-50 chars)",
    ),
    body: TrackRequest = ...,
    db: Session = Depends(get_session),
    _key: str = Depends(verify_api_key),
):
    """Track a shipment via carrier API and save/update in DB."""
    if body.project_id is not None:
        _validate_project_reference(db, body.project_id)
    result = track_and_save(
        tracking_number=tracking_number.upper(),
        recipient=body.recipient,
        items=body.shipment_name,
        destination=body.destination,
        show_date=body.show_date,
        exhibition_name=body.exhibition_name or "Unknown Exhibition",
        db=db,
        cs=body.cs,
        no_of_box=body.no_of_box,
        project_id=body.project_id,
    )
    if "error" in result:
        raise HTTPException(status_code=400, detail=result["error"])
    return result


# ---------------------------------------------------------------------------
# Batch import from Excel - uses BackgroundTasks so the response is immediate
# ---------------------------------------------------------------------------

def _process_excel_import(contents: bytes, db: Session):
    """Parse Excel rows and track each shipment, supporting Master/Child vertical nesting logic."""
    df = pd.read_excel(io.BytesIO(contents))
    # Normalize column names for easier lookup
    df.columns = [c.strip().lower().replace(" ", "_").replace("#", "").replace("/", "") for c in df.columns]

    # Required columns check (allowing for either tracking_number OR master/child structure)
    has_legacy = "tracking_number" in df.columns
    has_mps = "master_awb" in df.columns or "child_awb" in df.columns
    
    if not has_legacy and not has_mps:
        logger.error("Excel import: missing tracking columns (tracking_number or master_awb/child_awb)")
        return {"error": "Missing required tracking columns. Ensure 'Master AWB' or 'Tracking Number' exists.", "success": 0, "failed": 0}

    success = 0
    failed = 0
    errors = []
    
    last_master_awb = None
    
    for _, row in df.iterrows():
        tracking_num, master_to_use, is_master, last_master_awb = _resolve_tracking_row(
            master_awb=row.get("master_awb"),
            child_awb=row.get("child_awb"),
            legacy_tracking=row.get("tracking_number"),
            last_master_awb=last_master_awb,
        )
        if not tracking_num:
            continue

        # Metadata extraction
        items_name = str(row.get("name") or row.get("items") or "").strip() if pd.notna(row.get("name")) or pd.notna(row.get("items")) else None
        recipient = str(row.get("client_name") or row.get("recipient") or "").strip() if pd.notna(row.get("client_name")) or pd.notna(row.get("recipient")) else items_name
        show_date = str(row.get("show_date")).strip() if pd.notna(row.get("show_date")) else None
        
        exhibition_name = str(row.get("exhibition_name") or row.get("show_city") or "Unknown Exhibition").strip()
        cs = str(row.get("cs") or row.get("cs_type")).strip() if pd.notna(row.get("cs")) or pd.notna(row.get("cs_type")) else None
        no_of_box = str(row.get("no_of_box") or row.get("boxes")).strip() if pd.notna(row.get("no_of_box")) or pd.notna(row.get("boxes")) else None
        destination_hint = (
            str(row.get("ship_to_location") or row.get("destination") or "").strip()
            if pd.notna(row.get("ship_to_location")) or pd.notna(row.get("destination"))
            else None
        )
        
        project_id = int(row["project_id"]) if "project_id" in df.columns and pd.notna(row.get("project_id")) else None

        if project_id is not None and not db.get(DashboardProject, project_id):
            failed += 1
            errors.append(f"{tracking_num}: linked project not found")
            continue

        res = track_and_save(
            tracking_number=tracking_num.upper(),
            recipient=recipient,
            items=items_name,
            show_date=show_date,
            exhibition_name=exhibition_name,
            db=db,
            cs=cs,
            no_of_box=no_of_box,
            project_id=project_id,
            destination=destination_hint,
            # Pass MPS flags
            master_tracking_number=master_to_use,
            is_master=is_master,
            remarks=str(row.get("remarks")).strip() if pd.notna(row.get("remarks")) else None,
            booking_date=str(row.get("booking_dt")).strip() if pd.notna(row.get("booking_dt")) else None
        )
        
        if "error" in res:
            logger.warning("Import failed for %s: %s", tracking_num, res["error"])
            failed += 1
            errors.append(f"{tracking_num}: {res['error']}")
        else:
            success += 1

    logger.info("Excel import complete: %d succeeded, %d failed", success, failed)
    return {"success": success, "failed": failed, "errors": errors}


@router.post("/import-excel", status_code=200)
async def import_excel(
    file: UploadFile = File(...),
    db: Session = Depends(get_session),
    _key: str = Depends(verify_api_key),
):
    """
    Import shipments from an Excel file (.xlsx/.xls).
    Processing runs in a threadpool to prevent blocking the async event loop.
    Expected columns: tracking_number, name (optional), show_date (optional)
    """
    if not (file.filename or "").lower().endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="Invalid file format. Upload an .xlsx or .xls file.")

    contents = await file.read(MAX_EXCEL_FILE_SIZE + 1)
    if len(contents) > MAX_EXCEL_FILE_SIZE:
        raise HTTPException(status_code=413, detail="File too large. Maximum allowed size is 5 MB.")

    result = await run_in_threadpool(_process_excel_import, contents, db)
    
    if "error" in result:
        raise HTTPException(status_code=400, detail=result["error"])
        
    return {
        "status": "completed",
        "success": result["success"],
        "failed": result["failed"],
        "errors": result["errors"],
        "message": f"Successfully imported {result['success']} shipments. {result['failed']} failed.",
    }


def _clean_sheet_value(value: Optional[str]) -> Optional[str]:
    token = str(value or "").strip()
    return None if token.lower() in EMPTY_TRACKING_TOKENS else token


def _sheet_row_has_data(row: SheetRow) -> bool:
    return any(
        _clean_sheet_value(getattr(row, field))
        for field in (
            "ship_to_location",
            "client_name",
            "booking_date",
            "show_date",
            "show_city",
            "cs_type",
            "no_of_box",
            "courier",
            "master_awb",
            "child_awb",
            "remarks",
        )
    )


def _looks_like_dhl(courier: Optional[str]) -> bool:
    return "dhl" in str(courier or "").strip().lower()


def _resolve_sheet_carrier(tracking_number: str, courier: Optional[str]) -> str:
    detected = detect_carrier(tracking_number)
    if detected != "Unknown":
        return detected

    courier_label = str(courier or "").strip().lower()
    if "fedex" in courier_label or "fed ex" in courier_label:
        return "FedEx" if detected == "FedEx" else "Unknown"
    if "dhl" in courier_label:
        return "DHL" if detected == "DHL" else "Unknown"
    if "ups" in courier_label:
        return "UPS" if detected == "UPS" else "Unknown"
    return "Unknown"


def _sheet_failure(
    row: SheetRow,
    row_number: int,
    tracking_number: Optional[str],
    reason: str,
    *,
    carrier: Optional[str] = None,
) -> dict:
    return {
        "row_number": row_number,
        "tracking_number": tracking_number,
        "master_awb": _normalize_tracking_cell(row.master_awb),
        "child_awb": _normalize_tracking_cell(row.child_awb),
        "carrier": carrier or _resolve_sheet_carrier(tracking_number or "", row.courier),
        "reason": reason,
    }


def _upsert_sheet_shipment_metadata(
    *,
    db: Session,
    row: SheetRow,
    tracking_number: str,
    master_tracking_number: Optional[str],
    is_master: bool,
    carrier: str,
) -> dict:
    existing = db.exec(select(Shipment).where(Shipment.tracking_number == tracking_number)).first()
    now_iso = datetime.now(timezone.utc).isoformat()
    sync_event = {
        "description": "Shipment synced from Google Sheet",
        "location": _clean_sheet_value(row.ship_to_location) or "",
        "status": "Pending carrier refresh",
        "date": now_iso,
    }

    if existing:
        existing.carrier = carrier or existing.carrier
        existing.recipient = _clean_sheet_value(row.client_name) or existing.recipient
        existing.show_date = _clean_sheet_value(row.show_date) or existing.show_date
        existing.exhibition_name = _clean_sheet_value(row.show_city) or existing.exhibition_name
        existing.destination = _clean_sheet_value(row.ship_to_location) or existing.destination
        existing.booking_date = _clean_sheet_value(row.booking_date) or existing.booking_date
        existing.cs = _clean_sheet_value(row.cs_type) or existing.cs
        existing.cs_type = _clean_sheet_value(row.cs_type) or existing.cs_type
        existing.no_of_box = _clean_sheet_value(row.no_of_box) or existing.no_of_box
        existing.remarks = _clean_sheet_value(row.remarks) or existing.remarks
        existing.master_tracking_number = master_tracking_number
        existing.is_master = is_master
        if not existing.history:
            existing.history = [sync_event]
        db.add(existing)
        return {"tracking_number": tracking_number, "status": "success", "carrier": carrier, "outcome": "updated"}

    shipment = Shipment(
        tracking_number=tracking_number,
        carrier=carrier,
        status="Pending",
        recipient=_clean_sheet_value(row.client_name) or "",
        exhibition_name=_clean_sheet_value(row.show_city) or "Unknown Exhibition",
        show_date=_clean_sheet_value(row.show_date),
        origin="Unknown",
        destination=_clean_sheet_value(row.ship_to_location) or "Unknown",
        eta="TBD",
        progress=0,
        items="Child Package" if master_tracking_number else "Package",
        history=[sync_event],
        cs=_clean_sheet_value(row.cs_type),
        no_of_box=_clean_sheet_value(row.no_of_box),
        booking_date=_clean_sheet_value(row.booking_date),
        show_city=_clean_sheet_value(row.show_city),
        cs_type=_clean_sheet_value(row.cs_type),
        remarks=_clean_sheet_value(row.remarks),
        master_tracking_number=master_tracking_number,
        is_master=is_master,
        child_parcels=[],
    )
    db.add(shipment)
    return {"tracking_number": tracking_number, "status": "success", "carrier": carrier, "outcome": "created"}


def _process_webhook_payload(payload: WebhookPayload, db: Session):
    imported = 0
    failed = 0
    skipped = 0
    failures = []
    skipped_rows = []
    imported_rows = []
    seen_tracking_numbers: set[str] = set()
    last_master_awb = None

    for index, row in enumerate(payload.rows):
        row_number = row.row_number or index + 2
        tracking_number, master_to_use, is_master, last_master_awb = _resolve_tracking_row(
            master_awb=row.master_awb,
            child_awb=row.child_awb,
            last_master_awb=last_master_awb,
        )
        if not tracking_number:
            if _sheet_row_has_data(row):
                failed += 1
                failures.append(_sheet_failure(row, row_number, None, "Missing master_awb and child_awb"))
            else:
                skipped += 1
                skipped_rows.append({"row_number": row_number, "reason": "Blank row"})
            continue

        if tracking_number in seen_tracking_numbers:
            skipped += 1
            skipped_rows.append({
                "row_number": row_number,
                "tracking_number": tracking_number,
                "master_awb": _normalize_tracking_cell(row.master_awb),
                "child_awb": _normalize_tracking_cell(row.child_awb),
                "reason": "Duplicate tracking number in this sync payload",
            })
            continue
        seen_tracking_numbers.add(tracking_number)

        if not is_master and not master_to_use:
            failed += 1
            failures.append(_sheet_failure(
                row,
                row_number,
                tracking_number,
                "Child shipment is missing a master_awb relationship",
            ))
            continue

        carrier = _resolve_sheet_carrier(tracking_number, row.courier)
        if carrier == "Unknown":
            failed += 1
            reason = DHL_AWB_FORMAT_ERROR if _looks_like_dhl(row.courier) else "Unsupported carrier or tracking number format"
            failures.append(_sheet_failure(row, row_number, tracking_number, reason, carrier=carrier))
            continue

        try:
            if payload.track_live:
                res = track_and_save(
                    tracking_number=tracking_number,
                    recipient=row.client_name,
                    items=None,
                    show_date=row.show_date,
                    exhibition_name=row.show_city or "Unknown Exhibition",
                    db=db,
                    cs=row.cs_type,
                    no_of_box=row.no_of_box,
                    project_id=None,
                    destination=row.ship_to_location,
                    booking_date=row.booking_date,
                    show_city=row.show_city,
                    cs_type=row.cs_type,
                    remarks=row.remarks,
                    master_tracking_number=master_to_use,
                    is_master=is_master,
                )
                outcome = "tracked"
            else:
                res = _upsert_sheet_shipment_metadata(
                    db=db,
                    row=row,
                    tracking_number=tracking_number,
                    master_tracking_number=master_to_use,
                    is_master=is_master,
                    carrier=carrier,
                )
                outcome = res.get("outcome", "imported")

            if "error" in res:
                failed += 1
                failures.append(_sheet_failure(row, row_number, tracking_number, res["error"], carrier=carrier))
            else:
                imported += 1
                imported_rows.append({
                    "row_number": row_number,
                    "tracking_number": tracking_number,
                    "master_tracking_number": master_to_use,
                    "carrier": carrier,
                    "outcome": outcome,
                })
        except Exception as exc:
            logger.exception("Error processing Google Sheet row %s (%s)", row_number, tracking_number)
            failed += 1
            failures.append(_sheet_failure(row, row_number, tracking_number, str(exc), carrier=carrier))

    if not payload.track_live:
        db.commit()

    errors = [
        f"Row {failure['row_number']} ({failure.get('tracking_number') or 'no tracking'}): {failure['reason']}"
        for failure in failures
    ]

    return {
        "received": len(payload.rows),
        "processed": imported + failed + skipped,
        "imported": imported,
        "success": imported,
        "failed": failed,
        "skipped": skipped,
        "tracking_mode": "live" if payload.track_live else "metadata_only",
        "errors": errors,
        "failures": failures,
        "skipped_rows": skipped_rows,
        "imported_rows": imported_rows,
    }


@router.post("/webhook/google-sheet", status_code=200)
async def google_sheet_webhook(
    payload: WebhookPayload,
    db: Session = Depends(get_session),
    # Optional API key for Google Apps Script to authenticate
    _key: str = Depends(verify_api_key),
):
    """
    Webhook to receive batch imports from Google Sheet.
    Implements Vertical Logic for Master/Child AWBs.
    """
    result = await run_in_threadpool(_process_webhook_payload, payload, db)
    return {
        "status": "completed",
        **result,
        "message": (
            f"Received {result['received']} row(s). Imported {result['imported']}, "
            f"failed {result['failed']}, skipped {result['skipped']}."
        ),
    }


@router.get("/export-excel")
def export_shipments(
    shipment_ids: Optional[str] = Query(
        default=None,
        description="Comma-separated shipment IDs to export. If omitted, exports all active shipments.",
    ),
    db: Session = Depends(get_session),
    _key: str = Depends(verify_api_key),
):
    """Export shipments to Excel with requested formatting."""
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from datetime import date, datetime, timedelta

    requested_ids: list[int] = []
    if shipment_ids:
        for token in shipment_ids.split(","):
            part = token.strip()
            if not part:
                continue
            if not part.isdigit():
                raise HTTPException(status_code=400, detail=f"Invalid shipment id: {part}")
            requested_ids.append(int(part))
        if not requested_ids:
            raise HTTPException(status_code=400, detail="No valid shipment ids provided for export")

    # Fetch non-archived shipments; optionally scoped to requested ids.
    statement = select(Shipment).where(Shipment.is_archived == False)
    if requested_ids:
        statement = statement.where(Shipment.id.in_(requested_ids))
    shipments = db.exec(statement).all()
    if requested_ids:
        order = {sid: idx for idx, sid in enumerate(requested_ids)}
        shipments.sort(key=lambda s: order.get(s.id or 0, len(order)))
    
    # Create Workbook
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Shipments"

    today_str = datetime.now().strftime("%d.%m.%Y")
    
    headers = [
        "Ship to location", "Client Name", "Booking dt.", "Show date", 
        "Show City", "C/S", "No of Box", "Courier", "Master AWB", 
        "Child AWB #", "Current Status", "Remarks", "Last Scan date / Same place"
    ]
    
    # Styles
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    alignment_center = Alignment(horizontal="center", vertical="center")
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    # Write headers
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = alignment_center

    def _safe_date(raw_value):
        if not raw_value:
            return ""
        try:
            dt = datetime.fromisoformat(str(raw_value).replace("Z", "+00:00"))
            return dt.strftime("%d.%m.%Y")
        except (ValueError, TypeError):
            return str(raw_value)[:10]

    def _safe_time(raw_value):
        if not raw_value:
            return ""
        token = str(raw_value).strip()
        if ":" not in token and "T" not in token:
            return ""
        try:
            dt = datetime.fromisoformat(token.replace("Z", "+00:00"))
            return dt.strftime("%I:%M %p")
        except (ValueError, TypeError):
            try:
                parsed = pd.to_datetime(token, errors="coerce", dayfirst=True)
                if pd.isna(parsed):
                    return ""
                return parsed.strftime("%I:%M %p")
            except Exception:
                return ""

    def _format_booking_date(raw_value):
        formatted = _safe_date(raw_value)
        return formatted if formatted else ""

    def _status_date_time_location(raw_dt, raw_location):
        date_part = _safe_date(raw_dt)
        time_part = _safe_time(raw_dt)
        location = str(raw_location or "").strip()

        datetime_label = " ".join(token for token in [date_part, time_part] if token).strip()
        parts = [token for token in [datetime_label, location] if token]
        return " | ".join(parts) if parts else dash

    def _build_master_latest(shipment: Shipment):
        if shipment.history:
            latest = shipment.history[0]
            return _status_date_time_location(latest.get("date"), latest.get("location")), _safe_date(latest.get("date"))
        return _status_date_time_location(shipment.last_scan_date, shipment.destination or shipment.origin), _safe_date(shipment.last_scan_date)

    def _build_child_latest(parent: Shipment, child: dict):
        c_date_raw = child.get("last_date")
        c_loc = child.get("last_location") or ""
        if (not c_loc) and parent.history:
            master_latest = parent.history[0]
            c_loc = master_latest.get("location", c_loc)
            if not c_date_raw:
                c_date_raw = master_latest.get("date")

        return _status_date_time_location(c_date_raw, c_loc), _safe_date(c_date_raw)

    def _build_child_latest_from_shipment(child: Shipment):
        if child.history:
            latest = child.history[0]
            return _status_date_time_location(latest.get("date"), latest.get("location")), _safe_date(latest.get("date"))
        return _status_date_time_location(child.last_scan_date, child.destination or child.origin), _safe_date(child.last_scan_date)

    def _parse_show_date(raw_value) -> Optional[date]:
        if raw_value is None:
            return None
        if isinstance(raw_value, datetime):
            return raw_value.date()
        if isinstance(raw_value, date):
            return raw_value

        token = str(raw_value).strip()
        if not token or token.lower() in EMPTY_TRACKING_TOKENS:
            return None

        parsed = pd.to_datetime(token, errors="coerce")
        if pd.isna(parsed):
            parsed = pd.to_datetime(token, errors="coerce", dayfirst=True)
        if pd.isna(parsed):
            return None
        return parsed.date()

    def _should_highlight_show_date(raw_value) -> bool:
        parsed = _parse_show_date(raw_value)
        if not parsed:
            return False
        today = date.today()
        window_end = today + timedelta(days=20)
        return today <= parsed <= window_end

    def _write_row(row_idx: int, values: list, bold_master_awb: bool = False):
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.border = border
            if headers[col - 1] == "Show date" and _should_highlight_show_date(val):
                cell.fill = yellow_fill
            if col in [3, 8]:
                cell.alignment = alignment_center
            if bold_master_awb and col == 9:
                cell.font = Font(bold=True)

    # Group records by linkage: rows with master_tracking_number are children.
    children_by_master = {}
    top_level_shipments = []
    for shipment in shipments:
        tn = (shipment.tracking_number or "").strip().upper()
        master_tn = (shipment.master_tracking_number or "").strip().upper()
        if master_tn and master_tn != tn:
            children_by_master.setdefault(master_tn, []).append(shipment)
        else:
            top_level_shipments.append(shipment)

    current_row = 2
    rendered_master_tns = set()
    dash = "-"

    for s in top_level_shipments:
        master_tn = (s.tracking_number or "").strip().upper()
        if master_tn:
            rendered_master_tns.add(master_tn)

        latest_status, h_date = _build_master_latest(s)
        row_data = [
            s.destination or dash,
            s.recipient or dash,
            _format_booking_date(s.booking_date) or dash,
            s.show_date or dash,
            s.show_city or s.exhibition_name or dash,
            s.cs_type or s.cs or dash,
            s.no_of_box or dash,
            s.carrier.upper() if s.carrier else dash,
            s.tracking_number,
            "",  # Child AWB is empty on master row
            latest_status,
            s.remarks or dash,
            h_date or dash,
        ]
        _write_row(current_row, row_data, bold_master_awb=True)
        current_row += 1

        rendered_child_tns = set()

        # Render child records saved as individual Shipment rows.
        for child in children_by_master.get(master_tn, []):
            child_status, child_date = _build_child_latest_from_shipment(child)
            child_tn = (child.tracking_number or "").strip().upper()
            if child_tn:
                rendered_child_tns.add(child_tn)

            child_data = [
                child.destination or s.destination or dash,
                child.recipient or s.recipient or dash,
                _format_booking_date(child.booking_date) or _format_booking_date(s.booking_date) or dash,
                child.show_date or s.show_date or dash,
                child.show_city or child.exhibition_name or s.show_city or s.exhibition_name or dash,
                child.cs_type or child.cs or s.cs_type or s.cs or dash,
                "",  # No of Box empty for child rows in export format
                child.carrier.upper() if child.carrier else (s.carrier.upper() if s.carrier else dash),
                "",  # Master AWB empty for child row
                child.tracking_number or dash,
                child_status,
                child.remarks or s.remarks or dash,
                child_date or dash,
            ]
            _write_row(current_row, child_data)
            current_row += 1

        # Render legacy JSON child parcels, skipping duplicates already rendered.
        for c in s.child_parcels or []:
            c_tn = str(c.get("tracking_number") or "").strip().upper()
            if c_tn and c_tn in rendered_child_tns:
                continue
            c_latest, c_date = _build_child_latest(s, c)
            child_data = [
                s.destination or dash,
                s.recipient or dash,
                _format_booking_date(s.booking_date) or dash,
                s.show_date or dash,
                s.show_city or s.exhibition_name or dash,
                s.cs_type or s.cs or dash,
                "",  # No of Box empty for child
                s.carrier.upper() if s.carrier else dash,
                "",  # Master AWB empty for child
                c.get("tracking_number") or dash,
                c_latest,
                s.remarks or dash,
                c_date or dash,
            ]
            _write_row(current_row, child_data)
            current_row += 1

    # Preserve orphan child records even if the master row is missing in DB.
    for master_tn, orphan_children in children_by_master.items():
        if master_tn in rendered_master_tns:
            continue
        for child in orphan_children:
            child_status, child_date = _build_child_latest_from_shipment(child)
            orphan_row = [
                child.destination or dash,
                child.recipient or dash,
                _format_booking_date(child.booking_date) or dash,
                child.show_date or dash,
                child.show_city or child.exhibition_name or dash,
                child.cs_type or child.cs or dash,
                "",  # Keep child row format
                child.carrier.upper() if child.carrier else dash,
                master_tn,
                child.tracking_number or dash,
                child_status,
                child.remarks or dash,
                child_date or dash,
            ]
            _write_row(current_row, orphan_row)
            current_row += 1
    # Adjust column widths
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except (ValueError, TypeError):
                logger.warning("Failed to evaluate length for cell value: %s", cell.value)
                pass
        ws.column_dimensions[column].width = min(max_length + 2, 40)

    # Save to buffer
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    headers = {
        'Content-Disposition': f'attachment; filename="shipments_export_{today_str}.xlsx"'
    }
    return StreamingResponse(
        output,
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers=headers
    )


# ---------------------------------------------------------------------------
# Stats card counts
# ---------------------------------------------------------------------------

@router.get("/stats")
def shipment_stats(db: Session = Depends(get_session)):
    """Return counts by status for the dashboard stat cards (SQL aggregation)."""
    return get_stats(db)


@router.post("/refresh", status_code=200)
def refresh_shipments(
    body: RefreshRequest,
    db: Session = Depends(get_session),
    _key: str = Depends(verify_api_key),
):
    """Refresh saved shipments from their carriers and hydrate missing MPS child parcels."""
    return refresh_tracked_shipments(
        db=db,
        shipment_ids=body.shipment_ids,
        include_children=body.include_children,
    )


# ---------------------------------------------------------------------------
# List & detail
# ---------------------------------------------------------------------------

@router.get("/", response_model=List[ShipmentResponse])
def list_shipments(
    skip: int = 0,
    limit: int = 100,
    db: Session = Depends(get_session),
):
    """List active (non-archived) shipments."""
    shipments = db.exec(select(Shipment).where(Shipment.is_archived == False).offset(skip).limit(limit)).all()
    return [_serialize_shipment(db, shipment) for shipment in shipments]


@router.get("/archived", response_model=List[ShipmentResponse])
def list_archived_shipments(
    skip: int = 0,
    limit: int = 100,
    db: Session = Depends(get_session),
    _key: str = Depends(verify_api_key),
):
    """List archived shipments (Storage)."""
    shipments = db.exec(select(Shipment).where(Shipment.is_archived == True).offset(skip).limit(limit)).all()
    return [_serialize_shipment(db, shipment) for shipment in shipments]


@router.patch("/{shipment_id}/archive", response_model=ShipmentResponse)
def archive_shipment(
    shipment_id: int,
    db: Session = Depends(get_session),
    _key: str = Depends(verify_api_key),
):
    """Toggle the archive status of a shipment."""
    from app.services.shipment_service import toggle_archive
    updated = toggle_archive(shipment_id, db)
    if not updated:
        raise HTTPException(status_code=404, detail="Shipment not found")
    return _serialize_shipment(db, updated)


@router.get("/mps/{shipment_id}", response_model=MPSDetailResponse)
def get_mps_detail(
    shipment_id: int,
    db: Session = Depends(get_session),
):
    """
    Return full MPS detail for a master shipment: its own fields plus an
    aggregated summary of all child parcels and their individual statuses.
    Returns 404 if not found, 400 if the shipment is not an MPS master.
    """
    shipment = db.get(Shipment, shipment_id)
    if not shipment:
        raise HTTPException(status_code=404, detail="Shipment not found")
    if not shipment.is_master:
        raise HTTPException(
            status_code=400,
            detail="This shipment is not a Multi-Piece Shipment master.",
        )
    return MPSDetailResponse.from_shipment(shipment)


@router.get("/{shipment_id}", response_model=ShipmentResponse)
def get_shipment(shipment_id: int, db: Session = Depends(get_session)):
    shipment = db.get(Shipment, shipment_id)
    if not shipment:
        raise HTTPException(status_code=404, detail="Shipment not found")
    return _serialize_shipment(db, shipment)


@router.get("/project/{project_id}", response_model=List[ShipmentResponse])
def list_project_shipments(
    project_id: int,
    db: Session = Depends(get_session),
):
    shipments = db.exec(
        select(Shipment)
        .where(Shipment.project_id == project_id)
        .where(Shipment.is_archived == False)
    ).all()
    return [_serialize_shipment(db, shipment) for shipment in shipments]


# ---------------------------------------------------------------------------
# Delete
# ---------------------------------------------------------------------------

@router.delete("/{shipment_id}", status_code=200)
def delete_shipment(
    shipment_id: int,
    db: Session = Depends(get_session),
    _key: str = Depends(verify_api_key),
):
    from app.services.shipment_service import batch_delete

    result = batch_delete([shipment_id], db)
    if result.get("count", 0) == 0:
        raise HTTPException(status_code=404, detail="Shipment not found")
    logger.info(
        "Deleted shipment id=%d with cascade count=%d",
        shipment_id,
        result.get("count", 0),
    )
    return {
        "message": "Shipment deleted successfully",
        "deleted_id": shipment_id,
        "deleted_count": result.get("count", 0),
        "deleted_ids": result.get("deleted_ids", []),
    }


# ---------------------------------------------------------------------------
# Batch Operations
# ---------------------------------------------------------------------------

@router.post("/batch/archive", status_code=200)
def batch_archive_shipments(
    body: BatchRequest,
    db: Session = Depends(get_session),
    _key: str = Depends(verify_api_key),
):
    """Batch update archive status for multiple shipments."""
    from app.services.shipment_service import batch_update_archive
    if body.archive is None:
        raise HTTPException(status_code=400, detail="Missing 'archive' boolean in request body")
    return batch_update_archive(body.shipment_ids, body.archive, db)


@router.post("/batch/delete", status_code=200)
def batch_delete_shipments(
    body: BatchRequest,
    db: Session = Depends(get_session),
    _key: str = Depends(verify_api_key),
):
    """Batch delete multiple shipments."""
    from app.services.shipment_service import batch_delete
    return batch_delete(body.shipment_ids, db)
