"""Integration tests for shipment API endpoints."""
from datetime import datetime, timedelta, timezone
import io

from openpyxl import load_workbook


def create_project(client, name="Shipment Project"):
    response = client.post(
        "/api/v1/projects/",
        json={
            "project_name": name,
            "stage": "Win",
            "board_stage": "TBC",
        },
    )
    assert response.status_code == 200, response.text
    return response.json()


class TestHealthEndpoints:
    def test_root(self, client):
        resp = client.get("/")
        assert resp.status_code == 200
        assert "running" in resp.json()["message"].lower()

    def test_health_check(self, client):
        resp = client.get("/api/health")
        assert resp.status_code == 200
        assert resp.json()["status"] == "ok"


class TestListShipments:
    def test_list_empty(self, client):
        resp = client.get("/api/v1/shipments/")
        assert resp.status_code == 200
        assert isinstance(resp.json(), list)

    def test_stats_empty(self, client):
        resp = client.get("/api/v1/shipments/stats")
        assert resp.status_code == 200
        data = resp.json()
        assert "total" in data
        assert "delivered" in data
        assert "transit" in data
        assert "exceptions" in data


class TestTrackShipment:
    def test_track_fedex_number(self, client):
        """Track a 12-digit FedEx number (mocked service in conftest)."""
        project = create_project(client, "FedEx Shipment Project")
        resp = client.post(
            "/api/v1/shipments/track/888598190302",
            json={"exhibition_name": "Test Exhibition", "project_id": project["id"]},
        )
        assert resp.status_code == 201
        data = resp.json()
        assert data["status"] == "success"
        assert data["carrier"] == "FedEx"
        assert data["tracking_number"] == "888598190302"

    def test_track_creates_db_record(self, client):
        """After tracking, shipment should appear in list."""
        project = create_project(client, "List Shipment Project")
        client.post(
            "/api/v1/shipments/track/999000111222",
            json={"exhibition_name": "Test Expo", "project_id": project["id"]},
        )
        resp = client.get("/api/v1/shipments/")
        tracking_numbers = [s["tracking_number"] for s in resp.json()]
        assert "999000111222" in tracking_numbers

    def test_track_invalid_short_number(self, client):
        """Short tracking number should fail validation."""
        resp = client.post("/api/v1/shipments/track/ABC", json={"exhibition_name": "Invalid Code"})
        assert resp.status_code == 422  # Pydantic validation error

    def test_track_invalid_special_chars(self, client):
        """Tracking number with special chars should fail regex validation."""
        resp = client.post("/api/v1/shipments/track/123-456-789", json={"exhibition_name": "SpChars"})
        assert resp.status_code == 422

    def test_track_unsupported_carrier(self, client):
        """Unknown format should return 400 with a helpful message."""
        project = create_project(client, "Unsupported Carrier Project")
        resp = client.post(
            "/api/v1/shipments/track/UNKNOWNCARRIER1",
            json={"exhibition_name": "No Format", "project_id": project["id"]},
        )
        assert resp.status_code == 400

    def test_track_marks_exception_when_no_movement_over_2_days(self, client, monkeypatch):
        project = create_project(client, "Stuck Shipment Project")
        stale_date = (datetime.now(timezone.utc) - timedelta(days=3, hours=1)).isoformat()

        def mock_stale_track(self, tn):
            return {
                "status": "In Transit",
                "origin": "Mumbai, IN",
                "destination": "Irving, TX, US",
                "eta": "2026-04-25",
                "progress": 40,
                "history": [
                    {
                        "description": "Arrived at facility",
                        "location": "Memphis, TN, US",
                        "status": "Arrived",
                        "date": stale_date,
                    }
                ],
            }

        monkeypatch.setattr("app.services.fedex.FedExService.track", mock_stale_track)

        resp = client.post(
            "/api/v1/shipments/track/888598190302",
            json={"exhibition_name": "Test Exhibition", "project_id": project["id"]},
        )
        assert resp.status_code == 201

        list_resp = client.get("/api/v1/shipments/")
        shipment = next(s for s in list_resp.json() if s["tracking_number"] == "888598190302")
        assert shipment["status"] == "Exception"
        assert shipment["progress"] == 10

    def test_track_marks_child_parcel_exception_when_stale(self, client, monkeypatch):
        project = create_project(client, "Stale Child Parcel Project")
        fresh_date = datetime.now(timezone.utc).isoformat()
        stale_child_date = (datetime.now(timezone.utc) - timedelta(days=4)).isoformat()

        def mock_mps_track(self, tn):
            return {
                "status": "In Transit",
                "origin": "Mumbai, IN",
                "destination": "Irving, TX, US",
                "eta": "2026-04-25",
                "progress": 40,
                "history": [
                    {
                        "description": "Processed",
                        "location": "Dallas, TX, US",
                        "status": "Processed",
                        "date": fresh_date,
                    }
                ],
                "is_master": True,
                "child_parcels": [
                    {
                        "tracking_number": "CHILD001",
                        "status": "In Transit",
                        "raw_status": "In Transit",
                        "origin": "Mumbai, IN",
                        "destination": "Irving, TX, US",
                        "eta": "2026-04-25",
                        "last_date": stale_child_date,
                        "last_location": "Memphis, TN, US",
                        "carrier": "FedEx",
                    }
                ],
            }

        monkeypatch.setattr("app.services.fedex.FedExService.track", mock_mps_track)

        resp = client.post(
            "/api/v1/shipments/track/888598190302",
            json={"exhibition_name": "Test Exhibition", "project_id": project["id"]},
        )
        assert resp.status_code == 201

        list_resp = client.get("/api/v1/shipments/")
        shipment = next(s for s in list_resp.json() if s["tracking_number"] == "888598190302")
        assert shipment["child_parcels"][0]["status"] == "Exception"

    def test_track_uses_entry_destination_when_carrier_destination_unknown(self, client, monkeypatch):
        project = create_project(client, "Destination Fallback Project")

        def mock_track_unknown_destination(self, tn):
            return {
                "status": "In Transit",
                "origin": "Mumbai, IN",
                "destination": "Unknown",
                "eta": "2026-04-25",
                "progress": 40,
                "history": [],
            }

        monkeypatch.setattr(
            "app.services.fedex.FedExService.track",
            mock_track_unknown_destination,
        )

        resp = client.post(
            "/api/v1/shipments/track/888598190302",
            json={
                "exhibition_name": "Test Exhibition",
                "project_id": project["id"],
                "destination": "Irving, TX, US",
            },
        )
        assert resp.status_code == 201

        list_resp = client.get("/api/v1/shipments/")
        assert list_resp.status_code == 200
        shipment = next(s for s in list_resp.json() if s["tracking_number"] == "888598190302")
        assert shipment["destination"] == "Irving, TX, US"


class TestGetShipment:
    def test_get_existing(self, client):
        project = create_project(client, "Get Shipment Project")
        client.post(
            "/api/v1/shipments/track/111111111111",
            json={"exhibition_name": "Get Exists", "project_id": project["id"]},
        )
        list_resp = client.get("/api/v1/shipments/")
        shipment = next(s for s in list_resp.json() if s["tracking_number"] == "111111111111")
        resp = client.get(f"/api/v1/shipments/{shipment['id']}")
        assert resp.status_code == 200
        assert resp.json()["tracking_number"] == "111111111111"

    def test_get_not_found(self, client):
        resp = client.get("/api/v1/shipments/99999")
        assert resp.status_code == 404


class TestDeleteShipment:
    def test_delete_existing(self, client):
        project = create_project(client, "Delete Shipment Project")
        client.post(
            "/api/v1/shipments/track/222222222222",
            json={"exhibition_name": "Del Exists", "project_id": project["id"]},
        )
        list_resp = client.get("/api/v1/shipments/")
        shipment = next(s for s in list_resp.json() if s["tracking_number"] == "222222222222")
        resp = client.delete(f"/api/v1/shipments/{shipment['id']}")
        assert resp.status_code == 200

    def test_delete_not_found(self, client):
        resp = client.delete("/api/v1/shipments/99999")
        assert resp.status_code == 404

    def test_deleted_shipment_not_in_list(self, client):
        project = create_project(client, "Delete Missing Shipment Project")
        client.post(
            "/api/v1/shipments/track/333333333333",
            json={"exhibition_name": "Del Missing", "project_id": project["id"]},
        )
        list_resp = client.get("/api/v1/shipments/")
        shipment = next(s for s in list_resp.json() if s["tracking_number"] == "333333333333")
        client.delete(f"/api/v1/shipments/{shipment['id']}")
        list_resp2 = client.get("/api/v1/shipments/")
        tracking_numbers = [s["tracking_number"] for s in list_resp2.json()]
        assert "333333333333" not in tracking_numbers


class TestExportShipment:
    def test_export_scopes_to_requested_shipment_ids(self, client):
        project = create_project(client, "Export Scope Project")
        client.post(
            "/api/v1/shipments/track/444444444444",
            json={"exhibition_name": "Export One", "project_id": project["id"]},
        )
        client.post(
            "/api/v1/shipments/track/555555555555",
            json={"exhibition_name": "Export Two", "project_id": project["id"]},
        )

        list_resp = client.get("/api/v1/shipments/")
        rows = list_resp.json()
        first = next(r for r in rows if r["tracking_number"] == "444444444444")

        export_resp = client.get(f"/api/v1/shipments/export-excel?shipment_ids={first['id']}")
        assert export_resp.status_code == 200

        wb = load_workbook(io.BytesIO(export_resp.content))
        ws = wb.active
        exported_tracking = {
            str(ws.cell(row=i, column=9).value).strip()
            for i in range(2, ws.max_row + 1)
            if ws.cell(row=i, column=9).value
        }
        assert "444444444444" in exported_tracking
        assert "555555555555" not in exported_tracking

    def test_export_highlights_show_date_only_within_next_20_days(self, client):
        project = create_project(client, "Export Highlight Project")
        in_window = (datetime.now().date() + timedelta(days=5)).isoformat()
        out_window = (datetime.now().date() + timedelta(days=30)).isoformat()

        client.post(
            "/api/v1/shipments/track/666666666666",
            json={
                "exhibition_name": "Show Soon",
                "project_id": project["id"],
                "show_date": in_window,
            },
        )
        client.post(
            "/api/v1/shipments/track/777777777779",
            json={
                "exhibition_name": "Show Later",
                "project_id": project["id"],
                "show_date": out_window,
            },
        )

        export_resp = client.get("/api/v1/shipments/export-excel")
        assert export_resp.status_code == 200

        wb = load_workbook(io.BytesIO(export_resp.content))
        ws = wb.active

        row_by_tracking = {}
        for i in range(2, ws.max_row + 1):
            tracking = ws.cell(row=i, column=9).value
            if tracking:
                row_by_tracking[str(tracking).strip()] = i

        def is_yellow(cell) -> bool:
            color = (cell.fill.start_color.rgb or cell.fill.fgColor.rgb or "").upper()
            return cell.fill.fill_type == "solid" and color in {"00FFFF00", "FFFFFF00", "FFFF00"}

        soon_row = row_by_tracking["666666666666"]
        later_row = row_by_tracking["777777777779"]
        assert is_yellow(ws.cell(row=soon_row, column=4)) is True
        assert is_yellow(ws.cell(row=later_row, column=4)) is False


class TestGoogleSheetWebhook:
    def test_webhook_treats_row_with_master_and_child_as_child_record(self, client):
        payload = {
            "rows": [
                {
                    "master_awb": "777777777777",
                    "child_awb": "",
                    "ship_to_location": "Dallas, TX, US",
                },
                {
                    # Google Sheet style: master is repeated and child has actual piece AWB.
                    "master_awb": "777777777777",
                    "child_awb": "777777777778",
                    "ship_to_location": "Dallas, TX, US",
                },
            ]
        }

        resp = client.post("/api/v1/shipments/webhook/google-sheet", json=payload)
        assert resp.status_code == 200
        assert resp.json()["failed"] == 0

        list_resp = client.get("/api/v1/shipments/")
        rows = list_resp.json()
        master = next(r for r in rows if r["tracking_number"] == "777777777777")
        child = next(r for r in rows if r["tracking_number"] == "777777777778")

        assert master["is_master"] is True
        assert child["is_master"] is False
        assert child["master_tracking_number"] == "777777777777"
